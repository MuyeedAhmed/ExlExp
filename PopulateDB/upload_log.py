import openpyxl
import json
import os
import subprocess

VERIFICATION_EXCEL_PATH = '/Users/muyeedahmed/Desktop/Gitcode/ExlExp/Imported_Data_For_Verification.xlsx'
DB_JSON_PATH = '/Users/muyeedahmed/Desktop/Gitcode/ExlExp/db.json'
MIGRATE_SCRIPT_PATH = '/Users/muyeedahmed/Desktop/Gitcode/ExlExp/migrate.js'

def main():
    if not os.path.exists(VERIFICATION_EXCEL_PATH):
        print(f"Error: Verification Excel file not found at: {VERIFICATION_EXCEL_PATH}")
        print("Please run `python3 PopulateDB/import_log.py` first to generate it.")
        return
        
    print(f"Loading verification workbook from {VERIFICATION_EXCEL_PATH}...")
    wb = openpyxl.load_workbook(VERIFICATION_EXCEL_PATH, data_only=True)
    
    # 1. Parse Accounts sheet
    ws_accounts = wb["Accounts"]
    cards = []
    
    # Read rows skipping header
    for row in ws_accounts.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        c_id, c_name, c_type = row[0], row[1], row[2]
        
        card_obj = {
            "id": c_id,
            "name": c_name
        }
        if c_type == "Checking":
            card_obj["isChecking"] = True
        elif c_type == "Saving":
            card_obj["isSaving"] = True
        elif c_type == "Brokerage":
            card_obj["isBrokerage"] = True
            
        cards.append(card_obj)
        
    # 2. Parse Transactions sheet
    ws_tx = wb["Transactions"]
    expenses = []
    
    for row in ws_tx.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
            
        # Headers: ["ID", "Date", "Account Name", "Account ID", "Description/FromTo", "Details", "Amount", "Is Fee", "Is Reward", "Reward Type", "Reward Value", "Is Interest"]
        e_id = row[0]
        e_date = str(row[1]).split(" ")[0] if row[1] else "" # strip time if any
        c_id = row[3]
        desc = row[4] or ""
        details = row[5] or ""
        amount = float(row[6]) if row[6] is not None else 0.0
        is_fee = str(row[7]).upper() == "TRUE"
        is_reward = str(row[8]).upper() == "TRUE"
        reward_type = row[9] if row[9] else None
        
        try:
            reward_value = float(row[10]) if row[10] is not None and str(row[10]).strip() != "" and str(row[10]).strip() != "None" else None
        except ValueError:
            reward_value = None
            
        is_interest = str(row[11]).upper() == "TRUE"
        
        exp_obj = {
            "id": e_id,
            "date": e_date,
            "creditCardId": c_id,
            "description": desc,
            "amount": amount,
            "isFee": is_fee,
            "isReward": is_reward,
        }
        if reward_type:
            exp_obj["rewardType"] = reward_type
        if reward_value is not None:
            exp_obj["rewardValue"] = reward_value
        if is_interest:
            exp_obj["isInterest"] = is_interest
            
        # If it has checking/savings fields, keep them
        if details:
            exp_obj["details"] = details
        if desc:
            exp_obj["fromTo"] = desc
            
        expenses.append(exp_obj)
        
    # Write to db.json
    db_data = {
        "cards": cards,
        "expenses": expenses
    }
    
    print(f"Writing parsed data to {DB_JSON_PATH}...")
    with open(DB_JSON_PATH, 'w') as f:
        json.dump(db_data, f, indent=2)
        
    print(f"Generated db.json with {len(cards)} accounts and {len(expenses)} transactions.")
    
    # Trigger cloud sync
    if os.path.exists(MIGRATE_SCRIPT_PATH):
        print(f"\nTriggering cloud sync to Supabase: {MIGRATE_SCRIPT_PATH}...")
        try:
            result = subprocess.run(['node', MIGRATE_SCRIPT_PATH], capture_output=True, text=True, check=True)
            print(result.stdout)
            print("--- Sync Completed Successfully! ---")
        except subprocess.CalledProcessError as e:
            print(f"Error running cloud sync migration:\n{e.stderr}")
    else:
        print(f"\nCloud sync script not found at: {MIGRATE_SCRIPT_PATH}")

if __name__ == '__main__':
    main()

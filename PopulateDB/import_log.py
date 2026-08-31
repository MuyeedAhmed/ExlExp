import openpyxl
import json
import os
import datetime
import random
from openpyxl.styles import Font, Alignment, PatternFill

# File paths
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Bank Log.xlsx')
VERIFICATION_EXCEL_PATH = '/Users/muyeedahmed/Desktop/Gitcode/ExlExp/PopulateDB/Imported_Data_For_Verification.xlsx'
DB_JSON_PATH = '/Users/muyeedahmed/Desktop/Gitcode/ExlExp/db.json'

def generate_id(prefix=""):
    chars = "abcdefghijklmnopqrstuvwxyz0123456789"
    rand = "".join(random.choice(chars) for _ in range(9))
    return f"{prefix}{rand}"

def clean_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def get_category(description, details="", is_transfer=False, amount=0.0):
    desc = str(description or '').lower().strip()
    det = str(details or '').lower().strip()
    
    if is_transfer:
        return 'Transfer'
        
    transfer_keywords = [
        'credit card bill pay', 'credit card payment', 'cc payment', 'cc bill pay', 'cc billpay',
        'transfer', 'payment received', 'payment - thank you', 'payment thank you', 'autopay', 'auto-pay',
        'payment to', 'from checking', 'to checking', 'from savings', 'to savings',
        'from chase', 'to chase', 'from santander', 'to santander', 'from sofi', 'to sofi',
        'from citizens', 'to citizens', 'from upgrade', 'to upgrade', 'from mspbna', 'to mspbna',
        'from robinhood', 'to robinhood', 'from schwab', 'to schwab', 'from webull', 'to webull',
        'from etoro', 'to etoro', 'from gemini', 'to gemini', 'from paypal', 'to paypal',
        'from venmo', 'to venmo', 'imported balance', 'calculated from transfer logs', 'zelle from paypal',
        'zelle to paypal', 'current balance'
    ]
    
    account_keywords = [
        'amex', 'chase', 'citi', 'boa', 'bofat', 'bofa', 'hood-g', 'delta-g', 'delta gold',
        'robinhood', 'schwab', 'webull', 'gemini', 'etoro', 'santander', 'sofi', 'upgrade',
        'citizens', 'mspbna', 'closed-chasesav', 'amex-hysa', 'closed-dsrv', 'closed-wf-aj',
        'savings', 'checking', 'saving', 'chultd', 'chsdt', 'chsap', 'chflx', 'freedom unlimited',
        'sapphire', 'freedom flex', 'gold card', 'blue cash'
    ]
    
    # Catch any transfer/payment indicating known accounts
    is_account_involved = any(acc in desc or acc in det for acc in account_keywords)
    has_transfer_action = any(act in desc or act in det for act in ['to ', 'from ', 'payment', 'bill', 'deposit', 'withdraw', 'transfer', 'zelle', 'split', 'pay', 'receive'])
    
    if is_account_involved and has_transfer_action:
        return 'Transfer'
        
    if any(desc == acc or det == acc for acc in account_keywords):
        return 'Transfer'
        
    if any(k in desc or k in det for k in transfer_keywords):
        return 'Transfer'
        
    rent_keywords = ['rent', 'landlord', '11 fer blvd', '11ferblvd', 'faiaz rent', 'mouinul rent']
    if any(k in desc or k in det for k in rent_keywords):
        return 'Rent'
        
    utility_keywords = [
        'utility', 'utilities', 'wifi', 'wi-fi', 'internet', 'electric', 'electricity', 
        'water bill', 'pse&g', 'pseg', 'comcast', 'xfinity', 'verizon', 'spectrum', 'optimum'
    ]
    if any(k in desc or k in det for k in utility_keywords):
        return 'Utilities (inc. WiFi)'
        
    car_keywords = ['car loan', 'car payment', 'car finance', 'honda financial', 'toyota financial', 'chase auto']
    if any(k in desc or k in det for k in car_keywords):
        return 'Car Payment'
        
    trans_keywords = [
        'gas', 'toll', 'e-zpass', 'ezpass', 'mta', 'subway', 'bus', 'uber', 'lyft', 'train', 
        'transit', 'parking', 'parking fee', 'american dream parking', 'acura rallye'
    ]
    if any(k in desc or k in det for k in trans_keywords) and not ('gas bill' in desc or 'gas bill' in det or 'electric & gas' in det):
        return 'Transportation (Gas/Toll)'
        
    grocery_keywords = [
        'grocery', 'groceries', 'supermarket', 'market', 'walmart', 'target', 'costco', 'aldi', 
        'kroger', 'whole foods', 'trader joe', 'apna bazar', '99 ranch', 'patel brothers', 
        'aarong', 'h-mart', 'hmart', 'shoprite', 'stop & shop', 'stop and shop', '7-eleven', '7 eleven'
    ]
    if any(k in desc or k in det for k in grocery_keywords):
        return 'Grocery'
        
    eat_keywords = [
        'restaurant', 'diner', 'cafe', 'coffee', 'starbucks', 'dunkin', 'pizza', 'burger', 
        'mcdonald', 'taco bell', 'sweetgreen', 'chipotle', 'eating out', 'food', 'bakery', 
        'umacha', 'mochi mochi', 'omusubi', 'taskin bakery', 'nathans', 'sprite', 'grill', 'sushi', 
        'boba', 'uber eats', 'doordash', 'grubhub', 'seamless', 'almas', 'dunkin donuts'
    ]
    if any(k in desc or k in det for k in eat_keywords):
        return 'Eating Out'
        
    necessary_keywords = [
        'laptop', 'computer', 'monitor', 'desk', 'keyboard', 'mouse', 'software', 'subscription', 
        'aws', 'google cloud', 'openai', 'anthropic', 'github', 'microsoft', 'apple', 'insurance', 
        'medical', 'dentist', 'doctor', 'copay', 'pharmacy', 'cvs', 'walgreens', 'textbook', 
        'njit', 'tuition', 'fee', 'annual fee', 'progressive', 'united health'
    ]
    if any(k in desc or k in det for k in necessary_keywords):
        return 'Necessary Others'
        
    lux_keywords = [
        'vacation', 'hotel', 'flight', 'airline', 'air fare', 'airbnb', 'travel', 'cruise', 
        'ticketmaster', 'concert', 'amc', 'movie', 'cinema', 'theatre', 'disney', 'universal studios', 
        'resort', 'vacation', 'trip', 'alaska air', 'united air', 'delta air', 'american airlines', 
        'jetblue', 'top gun', 'billiards', 'airalo', 'airgsm', 'hotel v'
    ]
    if any(k in desc or k in det for k in lux_keywords):
        return 'Lux Others'
        
    salary_keywords = ['salary', 'paycheck', 'payroll', 'direct deposit', 'account open bonus', 'account opening bonus', 'bank bonus']
    if any(k in desc or k in det for k in salary_keywords):
        return 'Salary'
        
    if 'zelle' in desc or 'zelle' in det:
        return 'Transfer'
        
    return 'Others'

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: {EXCEL_PATH} not found.")
        return

    print("Loading workbook (reading evaluated data)...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Define metadata mapping for sheets
    SHEETS_METADATA = {
        # Credit Cards
        'CitiSrt': {'id': 'card-citistrata', 'name': 'Citi Strata', 'type': 'cc'},
        'CitiDb': {'id': 'card-citidb', 'name': 'Citi Double Cash', 'type': 'cc'},
        'BoA': {'id': 'card-bofa', 'name': 'BofA Premium', 'type': 'cc'},
        'BoA-T': {'id': 'card-bofat', 'name': 'BofA Travel', 'type': 'cc'},
        'Hood-G': {'id': 'card-hoodg', 'name': 'Robinhood Gold', 'type': 'cc'},
        'Delta-G': {'id': 'card-deltag', 'name': 'Delta Gold', 'type': 'cc'},
        'AmEx-B': {'id': 'card-amexb', 'name': 'AmEx Blue', 'type': 'cc'},
        'AmEx-G': {'id': 'card-amexg', 'name': 'AmEx Gold', 'type': 'cc'},
        'ChSdt': {'id': 'card-chsdt', 'name': 'Chase Student', 'type': 'cc'},
        'ChUltd': {'id': 'card-chultd', 'name': 'Chase Freedom Unlimited', 'type': 'cc'},
        'ChSaP': {'id': 'card-chsap', 'name': 'Chase Sapphire', 'type': 'cc'},
        'ChFlx': {'id': 'card-chflx', 'name': 'Chase Freedom Flex', 'type': 'cc'},
        'Closed-Dsrv': {'id': 'card-closeddsrv', 'name': 'Closed-Dsrv', 'type': 'cc'},
        'Closed-WF-AJ': {'id': 'card-closedwfaj', 'name': 'Closed-WF-AJ', 'type': 'cc'},
        
        # Checking Accounts
        'Chase': {'id': 'card-chase', 'name': 'Chase Checking', 'type': 'checking'},
        'Santander': {'id': 'card-santander', 'name': 'Santander', 'type': 'checking'},
        'SoFi': {'id': 'card-sofi', 'name': 'SoFi', 'type': 'checking'},
        'Upgrade': {'id': 'card-upgrade', 'name': 'Upgrade', 'type': 'checking'},
        'Citizens': {'id': 'card-citizens', 'name': 'Citizens', 'type': 'checking'},
        
        # Savings Accounts
        'MSPBNA': {'id': 'card-mspbna', 'name': 'MSPBNA', 'type': 'saving'},
        'Closed-ChaseSav': {'id': 'card-closedchasesav', 'name': 'Closed-ChaseSav', 'type': 'saving'},
        'AmEx-HYSA': {'id': 'card-amexhysa', 'name': 'AmEx-HYSA', 'type': 'saving'},
    }
    
    # Brokerage accounts to register and calculate
    BROKERAGES = {
        'Robinhood': {'id': 'card-robinhood', 'name': 'Robinhood'},
        'Schwab': {'id': 'card-schwab', 'name': 'Schwab'},
        'Webull': {'id': 'card-webull', 'name': 'Webull'},
        'Gemini': {'id': 'card-gemini', 'name': 'Gemini'},
        'Etoro': {'id': 'card-etoro', 'name': 'Etoro'},
    }
    
    # Tracks net balance for brokerages based on checking/savings transfers
    brokerage_balances = {name: 0.0 for name in BROKERAGES}
    
    cards = []
    expenses = []
    
    # Register credit card, checking, and savings accounts in the db
    for sheet_name, meta in SHEETS_METADATA.items():
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet {sheet_name} not found in workbook.")
            continue
            
        card_obj = {
            "id": meta['id'],
            "name": meta['name']
        }
        if meta['type'] == 'checking':
            card_obj["isChecking"] = True
        elif meta['type'] == 'saving':
            card_obj["isSaving"] = True
            
        cards.append(card_obj)
        
    # Register brokerage accounts
    for b_name, meta in BROKERAGES.items():
        cards.append({
            "id": meta['id'],
            "name": meta['name'],
            "isBrokerage": True
        })
        
    # Helper to check for brokerage transfers
    def track_brokerage_transfer(from_val, details_val, amount, is_deposit):
        from_val_lower = str(from_val or '').lower()
        details_val_lower = str(details_val or '').lower()
        
        for b_name in BROKERAGES:
            # Match brokerage name
            if b_name == 'Robinhood':
                # Exclude Robinhood Gold CC transfers
                if 'robinhood gold' in from_val_lower or 'robinhood gold' in details_val_lower:
                    continue
                if 'robinhood' not in from_val_lower and 'robinhood' not in details_val_lower:
                    continue
            else:
                if b_name.lower() not in from_val_lower and b_name.lower() not in details_val_lower:
                    continue
            
            # If it's a deposit into checking/saving, it was a withdrawal from brokerage
            # If it's a withdrawal from checking/saving, it was a deposit into brokerage
            if is_deposit:
                brokerage_balances[b_name] -= amount
            else:
                brokerage_balances[b_name] += amount
                
    # 1. Parse Credit Card sheets (Fee column is completely ignored)
    def parse_card_sheet(sheet_name, card_id):
        sheet = wb[sheet_name]
        print(f"Parsing CC sheet: {sheet_name}...")
        
        rows = list(sheet.iter_rows(values_only=True))
        header_row_idx = -1
        col_indices = {}
        
        for r_idx, row in enumerate(rows):
            row_str = [str(x).lower().strip() if x is not None else "" for x in row]
            if 'date' in row_str and ('to' in row_str or 'from' in row_str) and ('spend' in row_str or 'withdraw' in row_str):
                header_row_idx = r_idx
                for c_idx, val in enumerate(row_str):
                    if val == 'date' and 'date' not in col_indices:
                        col_indices['date'] = c_idx
                    elif val in ['to', 'from'] and 'to' not in col_indices:
                        col_indices['to'] = c_idx
                    elif val in ['return', 'paid'] and 'return' not in col_indices:
                        col_indices['return'] = c_idx
                    elif val in ['spend', 'withdraw'] and 'spend' not in col_indices:
                        col_indices['spend'] = c_idx
                    elif val in ['reward', 'rewards'] and 'rewards' not in col_indices:
                        col_indices['rewards'] = c_idx
                break
                
        if header_row_idx == -1:
            print(f"Warning: Could not find header row for CC sheet {sheet_name}.")
            return
            
        current_date = None
        count = 0
        
        for row in rows[header_row_idx + 1:]:
            date_cell = row[col_indices['date']] if 'date' in col_indices and col_indices['date'] < len(row) else None
            if date_cell is not None:
                if isinstance(date_cell, (datetime.datetime, datetime.date)):
                    current_date = date_cell.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_cell).strip()
                    try:
                        dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                        current_date = dt.strftime('%Y-%m-%d')
                    except ValueError:
                        pass
                        
            if current_date is None:
                continue
                
            to_cell = row[col_indices['to']] if 'to' in col_indices and col_indices['to'] < len(row) else None
            merchant = str(to_cell).strip() if to_cell is not None else ''
            if not merchant or merchant.isdigit():
                continue
                
            spend_cell = row[col_indices['spend']] if 'spend' in col_indices and col_indices['spend'] < len(row) else None
            return_cell = row[col_indices['return']] if 'return' in col_indices and col_indices['return'] < len(row) else None
            rewards_cell = row[col_indices['rewards']] if 'rewards' in col_indices and col_indices['rewards'] < len(row) else None
            
            spend_val = 0.0
            if spend_cell is not None:
                try:
                    spend_val = float(spend_cell)
                except ValueError:
                    pass
                    
            return_val = 0.0
            if return_cell is not None:
                try:
                    return_val = float(return_cell)
                except ValueError:
                    pass
                    
            reward_val = None
            is_reward = False
            reward_type = None
            if rewards_cell is not None:
                try:
                    val = float(rewards_cell)
                    if val > 0.0:
                        reward_val = val
                        is_reward = True
                        reward_type = 'cashback'
                except ValueError:
                    pass
            
            if spend_val == 0.0 and return_val == 0.0 and not is_reward:
                continue
                
            # Log separate spend transaction if spend is present
            if spend_val != 0.0:
                expenses.append({
                    "id": generate_id("exp-"),
                    "description": merchant,
                    "amount": round(spend_val, 2),
                    "creditCardId": card_id,
                    "date": current_date,
                    "category": get_category(merchant, '', False, spend_val),
                    "isReward": False,
                    "rewardType": None,
                    "rewardValue": None
                })
                count += 1
                
            # Log separate return transaction if return is present
            if return_val != 0.0:
                expenses.append({
                    "id": generate_id("exp-"),
                    "description": f"Refund: {merchant}" if spend_val != 0.0 else merchant,
                    "amount": round(-return_val if return_val > 0 else return_val, 2),
                    "creditCardId": card_id,
                    "date": current_date,
                    "category": get_category(merchant, '', False, -return_val),
                    "isReward": False,
                    "rewardType": None,
                    "rewardValue": None
                })
                count += 1
                
            # Log separate reward transaction if rewards are present
            if is_reward:
                expenses.append({
                    "id": generate_id("exp-"),
                    "description": f"Reward Credit: {merchant}" if spend_val == 0.0 and return_val == 0.0 else f"Reward Earned: {merchant}",
                    "amount": 0.0,  # reward transactions always have 0 amount (statement credits are handled in return column)
                    "creditCardId": card_id,
                    "date": current_date,
                    "category": get_category(merchant, '', False, 0.0),
                    "isReward": True,
                    "rewardType": reward_type,
                    "rewardValue": round(reward_val, 2)
                })
                count += 1
            
        print(f"Imported {count} transactions from sheet {sheet_name}.")
        
    # 2. Parse Checking sheets
    def parse_checking_sheet(sheet_name, card_id):
        sheet = wb[sheet_name]
        print(f"Parsing checking sheet: {sheet_name}...")
        
        rows = list(sheet.iter_rows(values_only=True))
        header_row_idx = -1
        col_indices = {}
        
        for r_idx, row in enumerate(rows):
            row_str = [str(x).lower().strip() if x is not None else "" for x in row]
            if 'date' in row_str and 'from' in row_str and ('withdraw' in row_str or 'withdraws' in row_str):
                header_row_idx = r_idx
                for c_idx, val in enumerate(row_str):
                    if val == 'date' and 'date' not in col_indices:
                        col_indices['date'] = c_idx
                    elif val == 'from' and 'from' not in col_indices:
                        col_indices['from'] = c_idx
                    elif val == 'salary' and 'salary' not in col_indices:
                        col_indices['salary'] = c_idx
                    elif val == 'deposit' and 'deposit' not in col_indices:
                        col_indices['deposit'] = c_idx
                    elif val in ['withdraw', 'withdraws'] and 'withdraw' not in col_indices:
                        col_indices['withdraw'] = c_idx
                    elif val == 'details' and 'details' not in col_indices:
                        col_indices['details'] = c_idx
                break
                
        if header_row_idx == -1:
            print(f"Warning: Could not find header row for checking sheet {sheet_name}.")
            return
            
        current_date = None
        count = 0
        
        for row in rows[header_row_idx + 1:]:
            # check if row is empty/only year indicator
            from_cell = row[col_indices.get('from')] if 'from' in col_indices else None
            details_cell = row[col_indices.get('details')] if 'details' in col_indices else None
            sal = row[col_indices.get('salary')] if 'salary' in col_indices else None
            dep = row[col_indices.get('deposit')] if 'deposit' in col_indices else None
            wd = row[col_indices.get('withdraw')] if 'withdraw' in col_indices else None
            
            if from_cell is None and details_cell is None and sal is None and dep is None and wd is None:
                continue
                
            date_cell = row[col_indices['date']] if 'date' in col_indices and col_indices['date'] < len(row) else None
            if date_cell is not None:
                if isinstance(date_cell, (datetime.datetime, datetime.date)):
                    current_date = date_cell.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_cell).strip()
                    try:
                        dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                        current_date = dt.strftime('%Y-%m-%d')
                    except ValueError:
                        pass
                        
            if current_date is None:
                continue
                
            from_to_val = str(from_cell).strip() if from_cell is not None else ''
            details_val = str(details_cell).strip() if details_cell is not None else ''
            
            amount = 0.0
            is_valid = False
            
            # Withdraw represents spending (negative expense)
            if wd is not None:
                try:
                    val = float(wd)
                    if val > 0:
                        amount = -val
                        is_valid = True
                        # Track brokerage transfer
                        track_brokerage_transfer(from_to_val, details_val, val, is_deposit=False)
                except ValueError:
                    pass
            
            # Salary/Deposit represents income (positive expense)
            is_salary_tx = False
            if not is_valid and sal is not None:
                try:
                    val = float(sal)
                    if val > 0:
                        amount = val
                        is_valid = True
                        is_salary_tx = True
                        # Track brokerage transfer
                        track_brokerage_transfer(from_to_val, details_val, val, is_deposit=True)
                except ValueError:
                    pass
                    
            if not is_valid and dep is not None:
                try:
                    val = float(dep)
                    if val > 0:
                        amount = val
                        is_valid = True
                        # Track brokerage transfer
                        track_brokerage_transfer(from_to_val, details_val, val, is_deposit=True)
                except ValueError:
                    pass
                    
            if not is_valid:
                continue
                
            expenses.append({
                "id": generate_id("exp-"),
                "description": from_to_val,
                "amount": round(amount, 2),
                "creditCardId": card_id,
                "date": current_date,
                "category": 'Salary' if is_salary_tx else get_category(from_to_val, details_val, False, amount),
                "fromTo": from_to_val,
                "details": details_val
            })
            count += 1
            
        print(f"Imported {count} transactions from checking sheet {sheet_name}.")

    # 3. Parse Savings sheets
    def parse_savings_sheet(sheet_name, card_id):
        sheet = wb[sheet_name]
        print(f"Parsing savings sheet: {sheet_name}...")
        
        rows = list(sheet.iter_rows(values_only=True))
        header_row_idx = -1
        col_indices = {}
        
        for r_idx, row in enumerate(rows):
            row_str = [str(x).lower().strip() if x is not None else "" for x in row]
            if 'date' in row_str and ('deposit' in row_str or 'withdraw' in row_str):
                header_row_idx = r_idx
                for c_idx, val in enumerate(row_str):
                    if val == 'date' and 'date' not in col_indices:
                        col_indices['date'] = c_idx
                    elif val == 'deposit' and 'deposit' not in col_indices:
                        col_indices['deposit'] = c_idx
                    elif val == 'withdraw' and 'withdraw' not in col_indices:
                        col_indices['withdraw'] = c_idx
                    elif val == 'interest' and 'interest' not in col_indices:
                        col_indices['interest'] = c_idx
                    elif val == 'details' and 'details' not in col_indices:
                        col_indices['details'] = c_idx
                break
                
        if header_row_idx == -1:
            print(f"Warning: Could not find header row for savings sheet {sheet_name}.")
            return
            
        current_date = None
        count = 0
        
        for row in rows[header_row_idx + 1:]:
            date_cell = row[col_indices['date']] if 'date' in col_indices and col_indices['date'] < len(row) else None
            if date_cell == 'Total':
                continue
                
            dep = row[col_indices.get('deposit')] if 'deposit' in col_indices else None
            wd = row[col_indices.get('withdraw')] if 'withdraw' in col_indices else None
            inte = row[col_indices.get('interest')] if 'interest' in col_indices else None
            details_cell = row[col_indices.get('details')] if 'details' in col_indices else None
            
            if dep is None and wd is None and inte is None and details_cell is None:
                continue
                
            if date_cell is not None:
                if isinstance(date_cell, (datetime.datetime, datetime.date)):
                    current_date = date_cell.strftime('%Y-%m-%d')
                else:
                    date_str = str(date_cell).strip()
                    try:
                        dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                        current_date = dt.strftime('%Y-%m-%d')
                    except ValueError:
                        pass
                        
            if current_date is None:
                continue
                
            details_val = str(details_cell).strip() if details_cell is not None else ''
            
            amount = 0.0
            is_valid = False
            is_interest = False
            
            # Withdraw is negative expense
            if wd is not None:
                try:
                    val = float(wd)
                    if val > 0:
                        amount = -val
                        is_valid = True
                        # Track brokerage transfer
                        track_brokerage_transfer(None, details_val, val, is_deposit=False)
                except ValueError:
                    pass
                    
            # Deposit is positive expense
            if not is_valid and dep is not None:
                try:
                    val = float(dep)
                    if val > 0:
                        amount = val
                        is_valid = True
                        # Track brokerage transfer
                        track_brokerage_transfer(None, details_val, val, is_deposit=True)
                except ValueError:
                    pass
                    
            # Interest is positive expense
            if not is_valid and inte is not None:
                try:
                    val = float(inte)
                    if val > 0:
                        amount = val
                        is_valid = True
                        is_interest = True
                except ValueError:
                    pass
                    
            if not is_valid:
                continue
                
            expenses.append({
                "id": generate_id("exp-"),
                "description": "Interest" if is_interest else details_val,
                "amount": round(amount, 2),
                "creditCardId": card_id,
                "date": current_date,
                "category": "Others" if is_interest else get_category(details_val, details_val, False, amount),
                "fromTo": "Interest" if is_interest else details_val,
                "details": details_val,
                "isInterest": is_interest
            })
            count += 1
            
        print(f"Imported {count} transactions from savings sheet {sheet_name}.")

    # Execute parsing
    for sheet_name, meta in SHEETS_METADATA.items():
        if sheet_name not in wb.sheetnames:
            continue
        if meta['type'] == 'cc':
            parse_card_sheet(sheet_name, meta['id'])
        elif meta['type'] == 'checking':
            parse_checking_sheet(sheet_name, meta['id'])
        elif meta['type'] == 'saving':
            parse_savings_sheet(sheet_name, meta['id'])
            
    # Write initial balance transactions for brokerage accounts
    print("\nCalculated Brokerage Balances:")
    for b_name, balance in brokerage_balances.items():
        print(f"  {b_name}: {balance:.2f}")
        meta = BROKERAGES[b_name]
        
        # Log a single transaction representing the current balance
        expenses.append({
            "id": generate_id("exp-"),
            "description": "Current Balance",
            "amount": round(balance, 2),
            "creditCardId": meta['id'],
            "date": "2026-08-30",  # current date
            "category": "Transfer",
            "fromTo": "Imported Balance",
            "details": "Calculated from transfer logs"
        })

    # Construct final db.json
    db_data = {
        "cards": cards,
        "expenses": expenses
    }
    
    print(f"\nWriting local backup database to {DB_JSON_PATH}...")
    with open(DB_JSON_PATH, 'w') as f:
        json.dump(db_data, f, indent=2)

    # 4. Create Verification Excel Workbook
    print(f"Generating verification Excel file at {VERIFICATION_EXCEL_PATH}...")
    v_wb = openpyxl.Workbook()
    
    # Setup Sheets
    ws_accounts = v_wb.active
    ws_accounts.title = "Accounts"
    ws_tx = v_wb.create_sheet(title="Transactions")
    
    # Styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal="center")
    
    # Write Accounts sheet
    accounts_headers = ["ID", "Name", "Type"]
    ws_accounts.append(accounts_headers)
    for col_idx in range(1, 4):
        cell = ws_accounts.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    for card in cards:
        c_type = "Credit Card"
        if card.get("isChecking"):
            c_type = "Checking"
        elif card.get("isSaving"):
            c_type = "Saving"
        elif card.get("isBrokerage"):
            c_type = "Brokerage"
        ws_accounts.append([card["id"], card["name"], c_type])
        
    # Autofit columns for Accounts
    for col in ws_accounts.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_accounts.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Write Transactions sheet (Fee column is excluded)
    tx_headers = [
        "ID", "Date", "Account Name", "Account ID", "Description/FromTo", "Details",
        "Category", "Amount", "Is Reward", "Reward Value", "Is Interest"
    ]
    ws_tx.append(tx_headers)
    for col_idx in range(1, len(tx_headers) + 1):
        cell = ws_tx.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        
    card_id_to_name = {c["id"]: c["name"] for c in cards}
    
    for exp in expenses:
        ws_tx.append([
            exp.get("id"),
            exp.get("date"),
            card_id_to_name.get(exp.get("creditCardId"), "Unknown Account"),
            exp.get("creditCardId"),
            exp.get("fromTo") or exp.get("description") or "",
            exp.get("details") or "",
            exp.get("category") or "Others",
            round(exp.get("amount", 0.0), 2),
            "TRUE" if exp.get("isReward") else "FALSE",
            round(exp.get("rewardValue"), 2) if exp.get("rewardValue") is not None else "",
            "TRUE" if exp.get("isInterest") else "FALSE"
        ])
        
    # Autofit columns for Transactions
    for col in ws_tx.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_tx.column_dimensions[col_letter].width = max(min(max_len + 3, 40), 10)
        
    v_wb.save(VERIFICATION_EXCEL_PATH)
    
    print(f"\nSuccess! Please inspect and verify the spreadsheet at: \n  {VERIFICATION_EXCEL_PATH}")
    print("When you're ready, run the following command to sync it to Supabase:")
    print("  python3 PopulateDB/upload_log.py")

if __name__ == '__main__':
    main()
